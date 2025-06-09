import pymysql
import pandas as pd
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import os
from dotenv import load_dotenv
import pymysql

load_dotenv()  # take environment variables from .env.




# Функция для получения квартала с использованием Tkinter
def get_quarter_input(callback, master_root):
    # создаём отдельное окно
    input_window = tk.Toplevel(master_root)
    input_window.title("Məlumat Bazasının Yenilənməsi")
    input_window.geometry("340x140+450+300")
    input_window.iconbitmap("Icon.ico")
    input_window.resizable(False, False)

    def on_submit(event=None):
        user_input = entry.get()
        try:
            quarter, year = user_input.split("/")
            quarter = int(quarter)
            year = int(year)
            last_quarter = f"{quarter}/{year}"

            if quarter == 1:
                prev_quarter = 4
                prev_year = year - 1
            else:
                prev_quarter = quarter - 1
                prev_year = year

            prev_quarter_str = f"{prev_quarter}/{prev_year}"

            input_window.destroy()
            callback(last_quarter, prev_quarter_str)

        except ValueError:
            messagebox.showerror("Xəta", "Zəhmət olmasa düzgün formatda daxil edin (rüb/il)")

    # Автоматически закрывать окно при закрытии главного
    def on_master_close():
        if input_window.winfo_exists():
            input_window.destroy()
        master_root.destroy()

    # привязываем обработчик закрытия главного окна
    master_root.protocol("WM_DELETE_WINDOW", on_master_close)

    # элементы окна
    label = tk.Label(input_window, text="Yoxlamaq istədiyiniz dövrü daxil edin (rüb/il)", font=("Arial", 12))
    label.pack(pady=10)

    entry = tk.Entry(input_window, font=("Arial", 14))
    entry.pack(pady=10)
    entry.focus()

    submit_button = tk.Button(input_window, text="Yenilə", font=("Arial", 12), command=on_submit)
    submit_button.pack(pady=10)

    entry.bind("<Return>", on_submit)




# Функция для обновления базы данных (ISP и PSTN)
def update_database_isp_pstn(last_quarter, prev_quarter_str):
    connection = None  # Инициализация переменной, чтобы избежать ошибки в блоке finally

    try:
        connection = pymysql.connect(
            host=os.getenv("DB_HOST"),
            user=os.getenv("DB_USER"),
            password=os.getenv("DB_PASSWORD"),
            database=os.getenv("DB_NAME")
        )

        # Проверка подключения
        if connection.open:
            print("Успешное подключение к базе данных")

            try:
                cursor = connection.cursor(pymysql.cursors.DictCursor)

                # Запрос для данных ISP
                sql_isp = f"""
                SELECT 
                    r.legal_name, 
                    e.voen,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.quarter END) AS quarter_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.ias END) AS ias_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.ias_xdsl END) AS ias_xdsl_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.ias_xpon END) AS ias_xpon_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.ias_axvi END) AS ias_axvi_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.si END) AS si_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.snuias END) AS snuias_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.snuias_msan_dslam END) AS snuias_msan_dslam_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.snuias_fttb END) AS snuias_fttb_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.ias_dysi END) AS ias_dysi_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.snuias_ftth END) AS snuias_ftth_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.snuias_other END) AS snuias_other_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.xvpgas END) AS xvpgas_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.isqt END) AS isqt_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.isust END) AS isust_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.iptas END) AS iptas_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.tvqs_itv END) AS tvqs_itv_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.xkibtcmdozs END) AS xkibtcmdozs_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.moafolu END) AS moafolu_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.moafols END) AS moafols_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.bdks END) AS bdks_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.igfols END) AS igfols_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.ivfols END) AS ivfols_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.msu END) AS msu_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.psu END) AS psu_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.moafolu_iolu END) AS moafolu_iolu_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.ivkkqu END) AS ivkkqu_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.mkkqrdiekttcku_kick END) AS mkkqrdiekttcku_kick_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.ku END) AS ku_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.ivfolu END) AS ivfolu_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.ivemplsps END) AS ivemplsps_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.bisqutakt END) AS bisqutakt_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.bisqutakt_giris END) AS bisqutakt_giris_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.bisqutakt_cixis END) AS bisqutakt_cixis_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.sgit END) AS sgit_prev,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.quarter END) AS quarter_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.ias END) AS ias_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.ias_xdsl END) AS ias_xdsl_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.ias_xpon END) AS ias_xpon_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.ias_axvi END) AS ias_axvi_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.si END) AS si_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.snuias END) AS snuias_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.snuias_msan_dslam END) AS snuias_msan_dslam_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.snuias_fttb END) AS snuias_fttb_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.ias_dysi END) AS ias_dysi_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.snuias_ftth END) AS snuias_ftth_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.snuias_other END) AS snuias_other_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.xvpgas END) AS xvpgas_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.isqt END) AS isqt_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.isust END) AS isust_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.iptas END) AS iptas_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.tvqs_itv END) AS tvqs_itv_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.xkibtcmdozs END) AS xkibtcmdozs_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.moafolu END) AS moafolu_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.moafols END) AS moafols_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.bdks END) AS bdks_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.igfols END) AS igfols_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.ivfols END) AS ivfols_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.msu END) AS msu_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.psu END) AS psu_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.moafolu_iolu END) AS moafolu_iolu_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.ivkkqu END) AS ivkkqu_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.mkkqrdiekttcku_kick END) AS mkkqrdiekttcku_kick_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.ku END) AS ku_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.ivfolu END) AS ivfolu_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.ivemplsps END) AS ivemplsps_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.bisqutakt END) AS bisqutakt_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.bisqutakt_giris END) AS bisqutakt_giris_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.bisqutakt_cixis END) AS bisqutakt_cixis_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.sgit END) AS sgit_last
                FROM economic_isp e
                LEFT JOIN registration r ON e.voen = r.voen
                WHERE e.quarter IN ('{last_quarter}', '{prev_quarter_str}')
                GROUP BY r.legal_name, e.voen
                """
                cursor.execute(sql_isp)
                results_isp = cursor.fetchall()
                df_isp = pd.DataFrame(results_isp)

                # === Путь к файлам ===
                file_path_isp_data = 'data ISP.xlsx'
                file_path_isp = 'ISP.xlsx'

                # Сохранение данных ISP в файл Excel (data ISP.xlsx)
                with pd.ExcelWriter(file_path_isp_data, engine='openpyxl') as writer:
                    df_isp.to_excel(writer, index=False, sheet_name='Economics ISP')
                print("Данные ISP успешно сохранены в файл 'data ISP.xlsx'")

                # Открытие data ISP и выбор листа Economics ISP
                wb_isp_data = load_workbook(file_path_isp_data, data_only=True)
                ws_isp_data_sheet = wb_isp_data['Economics ISP']

                # Открытие ISP.xlsx и выбор листа Input
                wb_isp = load_workbook(file_path_isp)
                ws_isp_input = wb_isp['Input']

                # Очистка листа Input
                for row in ws_isp_input.iter_rows(min_row=1, max_row=ws_isp_input.max_row, min_col=1,
                                                  max_col=ws_isp_input.max_column):
                    for cell in row:
                        cell.value = None

                # Копирование данных в Input
                for row in ws_isp_data_sheet.iter_rows(min_row=1, max_row=ws_isp_data_sheet.max_row, min_col=1,
                                                       max_col=ws_isp_data_sheet.max_column):
                    for cell in row:
                        ws_isp_input.cell(row=cell.row, column=cell.column, value=cell.value)

                # Преобразование ячеек в числа, кроме столбцов A, C и T
                for row in ws_isp_input.iter_rows(min_row=1, max_row=ws_isp_input.max_row, min_col=1,
                                                  max_col=ws_isp_input.max_column):
                    for cell in row:
                        if cell.column not in [1, 3, 20]:  # A, C, T
                            try:
                                cell.value = float(cell.value)
                            except (ValueError, TypeError):
                                pass

                # Сохраняем только файл ISP.xlsx
                wb_isp.save(file_path_isp)
                print("Данные успешно скопированы и преобразованы в файл 'ISP.xlsx', изменения сохранены.")



                # Запрос для данных PSTN
                sql_pstn = f"""
                SELECT 
                    r.legal_name, 
                    e.voen,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.quarter END) AS quarter_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.tas END) AS tas_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.tqs END) AS tqs_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.tsqt END) AS tsqt_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.tsust END) AS tsust_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.tts END) AS tts_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.stt END) AS stt_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.sbduct END) AS sbduct_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.sbdugt END) AS sbdugt_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.moafolu END) AS moafolu_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.moafols END) AS moafols_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.atsus END) AS atsus_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.ss END) AS ss_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.quyu END) AS quyu_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.ku END) AS ku_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.ts END) AS ts_prev,
                    MAX(CASE WHEN e.quarter = '{prev_quarter_str}' THEN e.tces END) AS tces_prev,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.quarter END) AS quarter_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.tas END) AS tas_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.tqs END) AS tqs_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.tsqt END) AS tsqt_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.tsust END) AS tsust_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.tts END) AS tts_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.stt END) AS stt_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.sbduct END) AS sbduct_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.sbdugt END) AS sbdugt_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.moafolu END) AS moafolu_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.moafols END) AS moafols_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.atsus END) AS atsus_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.ss END) AS ss_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.quyu END) AS quyu_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.ku END) AS ku_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.ts END) AS ts_last,
                    MAX(CASE WHEN e.quarter = '{last_quarter}' THEN e.tces END) AS tces_last
                FROM economic_pstn e
                LEFT JOIN registration r ON e.voen = r.voen
                WHERE e.quarter IN ('{last_quarter}', '{prev_quarter_str}')
                GROUP BY r.legal_name, e.voen
                """
                cursor.execute(sql_pstn)
                results_pstn = cursor.fetchall()
                df_pstn = pd.DataFrame(results_pstn)

                # Сохранение данных PSTN в файл
                file_path_pstn = 'data PSTN.xlsx'
                with pd.ExcelWriter(file_path_pstn, engine='openpyxl') as writer:
                    df_pstn.to_excel(writer, index=False, sheet_name='Economics PSTN')
                print("Данные PSTN успешно сохранены в файл 'data PSTN.xlsx'")



                # Открытие Excel файла для редактирования (PSTN)
                wb_pstn = load_workbook(file_path_pstn)
                ws_pstn = wb_pstn['Economics PSTN']

                # Поиск "0200004261" в столбце B и запись в соответствующую строку в столбце A
                for row in ws_pstn.iter_rows(min_row=2, max_row=ws_pstn.max_row, min_col=2, max_col=2):
                    if row[0].value == "0200004261":
                        ws_pstn.cell(row=row[0].row, column=1,
                                     value='"NAXÇIVAN MUXTAR RESPUBLİKASININ RƏQƏMSAL İNKİŞAF VƏ NƏQLİYYAT NAZİRLİYİ"')

                # Сохранение изменений в Excel для PSTN
                wb_pstn.save(file_path_pstn)

                print("Изменения успешно сохранены в файл 'data PSTN.xlsx'")

                # Открытие файла data PSTN (чтобы не сохранить изменения)
                file_path_pstn_data = 'data PSTN.xlsx'
                wb_pstn_data = load_workbook(file_path_pstn_data)
                ws_pstn_data = wb_pstn_data['Economics PSTN']

                # Открытие файла PSTN
                file_path_pstn = 'PSTN.xlsx'
                wb_pstn = load_workbook(file_path_pstn)
                ws_pstn = wb_pstn['Input']

                # Очищаем лист Input в файле PSTN перед вставкой данных
                for row in ws_pstn.iter_rows(min_row=1, max_row=ws_pstn.max_row, min_col=1, max_col=ws_pstn.max_column):
                    for cell in row:
                        cell.value = None

                # Копируем данные с листа Economics PSTN из файла data PSTN
                for row in ws_pstn_data.iter_rows(min_row=1, max_row=ws_pstn_data.max_row, min_col=1,
                                                  max_col=ws_pstn_data.max_column):
                    for cell in row:
                        # Вставляем данные в файл PSTN, в лист Input
                        ws_pstn.cell(row=cell.row, column=cell.column, value=cell.value)

                # Преобразуем все ячейки в числовые значения, кроме столбцов A, C и T
                for row in ws_pstn.iter_rows(min_row=1, max_row=ws_pstn.max_row, min_col=1, max_col=ws_pstn.max_column):
                    for cell in row:
                        if cell.column != 1 and cell.column != 3 and cell.column != 20:  # Проверяем столбцы A (1), C (3) и T (20)
                            try:
                                # Преобразуем в число, если это возможно
                                cell.value = float(cell.value)
                            except (ValueError, TypeError):
                                # Если преобразование не удалось (например, текст), оставляем значение как есть
                                pass

                # Сохранение изменений только в файл PSTN (data PSTN не сохраняем)
                wb_pstn.save(file_path_pstn)

                print("Данные успешно скопированы в файл 'PSTN.xlsx' и изменения сохранены.")


            finally:
                # Закрытие курсора, если он был создан
                if 'cursor' in locals():
                    cursor.close()

        else:
            messagebox.showinfo("Xəta", "Məlumat bazasına qoşulmaq mümkün deyil!")

    except Exception as e:
        messagebox.showinfo("Xəta", "Məlumat bazasına qoşulduqda xəta baş verir. IKTA şəbəkəsinə qoşulun!")
    finally:
        if connection and connection.open:
            connection.close()
            messagebox.showinfo("Məlumat", "Məlumat bazasının yenilənməsi uğurla tamamlandı")





def dcs_tool_isp():
    import os
    import tkinter as tk
    from tkinter import messagebox
    import win32com.client as win32
    import warnings
    from openpyxl import load_workbook
    from PIL import Image, ImageTk

    # Отключаем предупреждения
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    # Функция для окна ввода
    class InputDialog(tk.Toplevel):
        def __init__(self, parent):
            super().__init__(parent)
            self.title("ISP-nin adı")
            self.geometry("300x120")
            self.iconbitmap("Icon.ico")
            self.resizable(False,False)

            # Получаем размеры основного окна
            parent_width = parent.winfo_width()
            parent_height = parent.winfo_height()

            # Получаем размеры окна диалога
            dialog_width = 350
            dialog_height = 120

            # Вычисляем координаты верхнего левого угла окна диалога
            position_top = parent.winfo_rooty() + (parent_height // 2) - (dialog_height // 2)
            position_left = parent.winfo_rootx() + (parent_width // 2) - (dialog_width // 2)

            # Устанавливаем позицию окна
            self.geometry(f'{dialog_width}x{dialog_height}+{position_left}+{position_top}')

            def get_isp_date_from_excel():
                wb = load_workbook("ISP.xlsx", data_only=True)  # Открываем файл с вычисленными значениями
                sheet = wb["Input"]  # Выбираем лист "Output"
                return sheet["AL3"].value  # Возвращаем значение из ячейки E2

            tk.Label(self, text=f"Zəhmət olmasa {get_isp_date_from_excel()} rüb və il üçün ISP-nin adını daxil edin:", font=("Arial", 10)).pack(pady=10)
            self.entry = tk.Entry(self, width=37, font=("Arial", 12))
            self.entry.pack(pady=5)

            # Привязываем событие Enter
            self.entry.bind('<Return>', self.on_submit_event)

            tk.Button(self, text="Axtar", command=self.on_submit).pack(pady=10)

            self.result = None

        def on_submit(self):
            self.result = self.entry.get()
            self.destroy()

        def on_submit_event(self, event=None):  # Убираем использование параметра event
            self.on_submit()  # Вызываем метод on_submit, если нажата клавиша Enter

    # Функция для поиска провайдера
    def search_isp():
        excel = None
        try:
            # Открываем кастомное окно для ввода
            dialog = InputDialog(root)
            root.wait_window(dialog)
            isp_name = dialog.result

            if not isp_name:
                messagebox.showwarning("Xəbərdarlıq", "ISP-nin adı daxil olunmayıb!")
                return

            current_dir = os.getcwd()
            isp_file_path = os.path.join(current_dir, 'ISP.xlsx')

            if not os.path.exists(isp_file_path):
                messagebox.showerror("Xəta", "ISP.xlsx faylı tapılmadı")
                return

            excel = win32.Dispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False

            isp_workbook = excel.Workbooks.Open(isp_file_path)
            output_sheet = isp_workbook.Sheets("Index (Output)")  # Используем лист с названием "Index (Output)"

            found_rows = []  # Список для хранения строк с совпадениями

            # Ищем имя ISP в столбце B начиная с 4-й строки до строки 155, начиная со второго символа
            for row in range(4, 155):  # Ищем до строки 155
                cell_raw = output_sheet.Cells(row, 2).Value
                if cell_raw is None:
                    continue
                cell_value = str(cell_raw).lower().strip()  # Преобразуем в строку и удаляем пробелы

                # Печать значений для диагностики
                print(f"Проверяется строка {row}: {cell_value}")

                # Ищем, начинается ли строка с нужного имени ISP, начиная с второго символа
                if len(cell_value) > 1 and cell_value[1:].startswith(isp_name.lower().strip()):
                    print(f"Найдено совпадение на строке {row}: {cell_value}")  # Диагностика совпадения
                    found_rows.append(row)

            # Печать результата для проверки
            print("Найденные строки:", found_rows)

            if len(found_rows) == 1:  # Если найдено одно совпадение
                # Собираем данные из столбцов C до R (это столбцы 3-30)
                found_row = found_rows[0]
                isp_data = []
                index = 1  # Нумерация начинается с 1
                for col in range(3, 30):  # Столбцы от C до R
                    cell_value = output_sheet.Cells(found_row, col).Value
                    if cell_value:  # Пропускаем пустые ячейки
                        isp_data.append(f"{index}. {str(cell_value)}")  # Добавляем номер и значение
                        index += 1  # Увеличиваем номер для следующей строки

                # Получаем название из ячейки B в найденной строке (это будет заголовок)
                isp_name_header = str(output_sheet.Cells(found_row, 2).Value)

                # Объединяем данные в одну строку
                isp_info = "\n".join(isp_data)

                # Показываем результат с названием как заголовок
                show_custom_result(f"{isp_name_header}\n\n{isp_info}")
            elif len(found_rows) > 1:  # Если найдено несколько совпадений
                messagebox.showwarning("Xəbərdarlıq", "Uyğun nəticələr var. Başlıq üçün daha çox simvol daxil edin")
            else:
                messagebox.showwarning("Xəbərdarlıq", "Bu ISP-nin adı tapılmadı")

        except Exception as e:
            messagebox.showerror("Xəta", f"Xəta baş verdi: {str(e)}")

        finally:
            if excel:
                excel.Quit()

    # Функция для отображения результатов
    def show_custom_result(result_text):
        result_window = tk.Toplevel(root)
        result_window.title("ISP üzrə məlumat")
        result_window.geometry("1000x600+10+10")
        result_window.iconbitmap("Icon.ico")
        result_window.configure(bg="#FFFAFA")  # Цвет фона можно изменить на любой другой

        # Разделяем результат на строки
        result_lines = result_text.split("\n")

        # Заголовок с жирным шрифтом
        tk.Label(result_window, text=result_lines[0], font=("Arial", 14, "bold"), anchor="center", bg="#FFFAFA", wraplength=900).pack(
            pady=10)

        # Создаем текстовое поле для вывода текста, которое будет доступно для копирования
        text_widget = tk.Text(result_window, font=("Arial", 12), wrap="word", padx=20, pady=20, height=20, width=90,
                              bg="#FFFAFA")
        text_widget.pack(padx=20, pady=20)

        # Вставляем текст в текстовое поле
        text_widget.insert(tk.END, "\n".join(result_lines[1:]))

        # Отключаем возможность редактирования, но оставляем возможность копирования
        text_widget.config(state=tk.DISABLED)

    # Функция для генерации отчета с полупрозрачным фоном
    def generate_report_isp():
        import os
        from tkinter import messagebox, filedialog
        import win32com.client as win32

        try:
            # Получение текущей директории и пути к Excel-файлу
            current_dir = os.getcwd()
            isp_file_path = os.path.join(current_dir, 'ISP.xlsx')

            if not os.path.exists(isp_file_path):
                messagebox.showerror("Xəta", "ISP.xlsx faylı tapılmadı")
                return

            # Запуск Excel
            excel = win32.Dispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False

            # Открываем рабочую книгу и нужный лист
            wb = excel.Workbooks.Open(isp_file_path)
            sheet = wb.Sheets("Index (Output)")

            # Собираем данные: имя провайдера + критерии
            report_lines = []
            for row in range(4, 155):  # Строки 4–154
                isp_name = sheet.Cells(row, 2).Value  # Столбец B
                if isp_name:
                    criteria = []
                    for col in range(3, 24):  # Столбцы C до W (3–23)
                        cell_value = sheet.Cells(row, col).Value
                        if cell_value:
                            criteria.append(str(cell_value))
                    report_lines.append(f"{isp_name}:\n  " + "\n  ".join(criteria))

            wb.Close(SaveChanges=False)

            if not report_lines:
                messagebox.showinfo("Məlumat", "ISP faylında məlumat tapılmadı")
                return

            def get_isp_date_from_excel():
                wb = load_workbook("ISP.xlsx", data_only=True)  # Открываем файл с вычисленными значениями
                sheet = wb["Input"]  # Выбираем лист "Output"
                return sheet["AL3"].value  # Возвращаем значение из ячейки E2

            # Создание PDF-файла (временный путь)
            from fpdf import FPDF

            # Создаём кастомный класс с автоматическим фоном на каждую страницу
            class PDFWithBackground(FPDF):
                def __init__(self, background_image, *args, **kwargs):
                    super().__init__(*args, **kwargs)
                    self.background_image = background_image

                def header(self):
                    # Устанавливаем фон на каждую страницу
                    self.image(self.background_image, x=0, y=0, w=210, h=297)

            # Создаём PDF с фоном
            background_image = os.path.join(current_dir, "Background_reports.jpg")
            if not os.path.exists(background_image):
                messagebox.showerror("Xəta", "Fayl Background_reports.jpg tapılmadı")
                return

            pdf = PDFWithBackground(background_image=background_image, format='A4', unit='mm')
            pdf.set_auto_page_break(auto=True, margin=15)

            # Добавляем шрифты
            pdf.add_font('DejaVu', '', 'DejaVuSans.ttf', uni=True)
            pdf.add_font('DejaVu', 'B', 'DejaVuSans-Bold.ttf', uni=True)

            pdf.add_page()

            # Добавляем название отчета
            pdf.set_font('DejaVu', 'B', 16)
            pdf.multi_cell(0, 10,
                           f'İNTERNET PROVAYDERLƏRİN {get_isp_date_from_excel()} RÜB VƏ İL ÜÇÜN TƏQDIM ETDIYI MƏLUMATLARLA BAĞLI XƏTALAR VƏ SƏBƏBLƏRI HAQQINDA HESABAT.',
                           0, 'C')

            # Отступ перед первым блоком
            y_position = pdf.get_y() + 10
            pdf.set_xy(10, y_position)

            # Перебираем блоки отчёта
            for block in report_lines:
                lines = block.strip().split("\n")
                if not lines:
                    continue

                provider_name = lines[0]
                criteria_lines = lines[1:]

                pdf.set_xy(10, y_position)
                pdf.set_font('DejaVu', 'B', 12)
                pdf.multi_cell(0, 10, txt=provider_name)
                y_position = pdf.get_y()

                pdf.set_font('DejaVu', '', 9)
                for idx, line in enumerate(criteria_lines, start=1):
                    formatted_line = f"{idx}. {line.strip()}"
                    pdf.multi_cell(0, 8, txt=formatted_line)
                    y_position = pdf.get_y()

                # Линия-разделитель
                pdf.set_line_width(0.5)
                pdf.line(10, y_position, 200, y_position)

                y_position += 5

                # Принудительная проверка на заполнение страницы
                if y_position > 270:
                    pdf.add_page()
                    y_position = 10

            # Диалог сохранения
            save_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                title="Yadda saxlayın",
                filetypes=[("PDF files", "*.pdf")],
                initialdir=os.getcwd(),
                initialfile="Report for ISPs.pdf",
                confirmoverwrite=True
            )

            if save_path:
                pdf.output(save_path)
                messagebox.showinfo("Uğur", f"Hesabat yadda saxlanıldı")
            else:
                messagebox.showwarning("Xəbərdarlıq", "Hesabat yadda saxlanılmadı")

        except Exception as e:
            messagebox.showerror("Xəta", f"Xəta baş verdi: {str(e)}")

        finally:
            if 'excel' in locals():
                excel.Quit()

    def get_isp_date_from_excel():
        wb = load_workbook("ISP.xlsx", data_only=True)  # Открываем файл с вычисленными значениями
        sheet = wb["Input"]  # Выбираем лист "Output"
        return sheet["AL3"].value  # Возвращаем значение из ячейки E2


    root = tk.Tk()
    root.title("ISP-lərin yoxlanılması")
    root.iconbitmap("Icon.ico")
    root.geometry("400x250+400+170")
    root.resizable(False, False)

    # Загружаем изображение с помощью Pillow
    image = Image.open("Background.jpg")
    background_image = ImageTk.PhotoImage(image)

    # Canvas для фонового изображения
    canvas = tk.Canvas(root, width=450, height=250, highlightthickness=0)
    canvas.pack(fill="both", expand=True)

    # Устанавливаем фон
    canvas.create_image(0, 0, image=background_image, anchor="nw")

    # Добавляем заголовок (прозрачный фон)
    canvas.create_text(200, 50, text="ISP-lərin yoxlanılması", font=("Helvetica", 22, "bold"), fill="black")

    # Список с позициями для кнопок: (x, y)
    button_positions = [(200, 120), (200, 190)]  # x, y для каждой кнопки

    # Размеры кнопок
    button_width = 25
    button_height = 2

    # Кнопка для генерации отчета
    ggenerate_button = tk.Button(root, text=f"Hesabat yaratmaq (dövr: {get_isp_date_from_excel()})",
                                 font=("Arial", 14, "bold"),
                                 command=lambda: generate_report_isp(),
                                 bg="#40E0D0", fg="black", activebackground="#0000FF", activeforeground="white",
                                 relief="raised", width=button_width, height=button_height)
    canvas.create_window(button_positions[0][0], button_positions[0][1], window=ggenerate_button)

    # Кнопка для поиска провайдера
    search_button = tk.Button(root, text=f"ISP üzrə axtarış (dövr: {get_isp_date_from_excel()})",
                              font=("Arial", 14, "bold"),
                              command=lambda: search_isp(),
                              bg="#7FFFD4", fg="black", activebackground="#0000FF", activeforeground="white",
                              relief="raised", width=button_width, height=button_height)
    canvas.create_window(button_positions[1][0], button_positions[1][1], window=search_button)

    root.mainloop()
    dcstool_main()


def dcs_tool_pstn():
    import os
    import tkinter as tk
    from tkinter import messagebox
    import win32com.client as win32
    import warnings
    from openpyxl import load_workbook
    from PIL import Image, ImageTk

    # Отключаем предупреждения
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    # Функция для окна ввода
    class InputDialog(tk.Toplevel):
        def __init__(self, parent):
            super().__init__(parent)
            self.title("PSTN-nin adı")
            self.geometry("300x120")
            self.iconbitmap("Icon.ico")
            self.resizable(False,False)

            # Получаем размеры основного окна
            parent_width = parent.winfo_width()
            parent_height = parent.winfo_height()

            # Получаем размеры окна диалога
            dialog_width = 360
            dialog_height = 120

            # Вычисляем координаты верхнего левого угла окна диалога
            position_top = parent.winfo_rooty() + (parent_height // 2) - (dialog_height // 2)
            position_left = parent.winfo_rootx() + (parent_width // 2) - (dialog_width // 2)

            # Устанавливаем позицию окна
            self.geometry(f'{dialog_width}x{dialog_height}+{position_left}+{position_top}')

            def get_pstn_date_from_excel():
                wb = load_workbook("PSTN.xlsx", data_only=True)  # Открываем файл с вычисленными значениями
                sheet = wb["Input"]  # Выбираем лист "Output"
                return sheet["T2"].value  # Возвращаем значение из ячейки E2


            tk.Label(self, text=f"Zəhmət olmasa {get_pstn_date_from_excel()} rüb və il üçün PSTN-nin adını daxil edin:", font=("Arial", 10)).pack(pady=10)
            self.entry = tk.Entry(self, width=38, font=("Arial", 12))
            self.entry.pack(pady=5)

            # Привязываем событие Enter
            self.entry.bind('<Return>', self.on_submit_event)

            tk.Button(self, text="Axtar", command=self.on_submit).pack(pady=10)

            self.result = None

        def on_submit(self):
            self.result = self.entry.get()
            self.destroy()

        def on_submit_event(self, event=None):  # Убираем использование параметра event
            self.on_submit()  # Вызываем метод on_submit, если нажата клавиша Enter

    # Функция для поиска провайдера
    def search_pstn():
        excel = None
        try:
            # Открываем кастомное окно для ввода
            dialog = InputDialog(root)
            root.wait_window(dialog)
            pstn_name = dialog.result

            if not pstn_name:
                messagebox.showwarning("Xəbərdarlıq", "PSTN-nin adı daxil olunmayıb!")
                return

            current_dir = os.getcwd()
            pstn_file_path = os.path.join(current_dir, 'PSTN.xlsx')

            if not os.path.exists(pstn_file_path):
                messagebox.showerror("Xəta", "PSTN.xlsx faylı tapılmadı")
                return

            excel = win32.Dispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False

            pstn_workbook = excel.Workbooks.Open(pstn_file_path)
            output_sheet = pstn_workbook.Sheets("Index (Output)")  # Используем лист с названием "Index (Output)"

            found_rows = []  # Список для хранения строк с совпадениями

            # Ищем имя PSTN в столбце B начиная с 3-й строки до строки 15, начиная со второго символа
            for row in range(2, 17):  # Ищем до строки 15
                cell_raw = output_sheet.Cells(row, 2).Value
                if cell_raw is None:
                    continue
                cell_value = str(cell_raw).lower().strip()  # Преобразуем в строку и удаляем пробелы

                # Печать значений для диагностики
                print(f"Проверяется строка {row}: {cell_value}")

                # Ищем совпадение начиная со второго символа в строке
                if len(cell_value) > 1 and cell_value[1:].startswith(pstn_name.lower().strip()):
                    print(f"Найдено совпадение на строке {row}: {cell_value}")  # Диагностика совпадения
                    found_rows.append(row)

            # Печать результата для проверки
            print("Найденные строки:", found_rows)

            if len(found_rows) == 1:  # Если найдено одно совпадение
                # Собираем данные из столбцов C до R (это столбцы 3-30)
                found_row = found_rows[0]
                pstn_data = []
                index = 1  # Нумерация начинается с 1
                for col in range(3, 30):  # Столбцы от C до R
                    cell_value = output_sheet.Cells(found_row, col).Value
                    if cell_value:  # Пропускаем пустые ячейки
                        pstn_data.append(f"{index}. {str(cell_value)}")  # Добавляем номер и значение
                        index += 1  # Увеличиваем номер для следующей строки

                # Получаем название из ячейки B в найденной строке (это будет заголовок)
                pstn_name_header = str(output_sheet.Cells(found_row, 2).Value)

                # Объединяем данные в одну строку
                pstn_info = "\n".join(pstn_data)

                # Показываем результат с названием как заголовок
                show_custom_result(f"{pstn_name_header}\n\n{pstn_info}")
            elif len(found_rows) > 1:  # Если найдено несколько совпадений
                messagebox.showwarning("Xəbərdarlıq", "Uyğun nəticələr var. Başlıq üçün daha çox simvol daxil edin")
            else:
                messagebox.showwarning("Xəbərdarlıq", "Bu PSTN-nin adı tapılmadı")

        except Exception as e:
            messagebox.showerror("Xəta", f"Xəta baş verdi: {str(e)}")

        finally:
            if excel:
                excel.Quit()

    # Функция для отображения результатов
    def show_custom_result(result_text):
        result_window = tk.Toplevel(root)
        result_window.title("PSTN üzrə məlumat")
        result_window.geometry("1000x600+10+10")
        result_window.iconbitmap("Icon.ico")
        result_window.configure(bg="#FFFAFA")  # Цвет фона можно изменить на любой другой
        result_window.resizable(False,False)

        # Разделяем результат на строки
        result_lines = result_text.split("\n")

        # Заголовок с жирным шрифтом
        tk.Label(result_window, text=result_lines[0], font=("Arial", 14, "bold"), anchor="center", bg="#FFFAFA", wraplength=900).pack(
            pady=10)

        # Создаем текстовое поле для вывода текста, которое будет доступно для копирования
        text_widget = tk.Text(result_window, font=("Arial", 12), wrap="word", padx=20, pady=20, height=20, width=90,
                              bg="#FFFAFA")
        text_widget.pack(padx=20, pady=20)

        # Вставляем текст в текстовое поле
        text_widget.insert(tk.END, "\n".join(result_lines[1:]))

        # Отключаем возможность редактирования, но оставляем возможность копирования
        text_widget.config(state=tk.DISABLED)

    # Функция для генерации отчета
    def generate_report_pstn():
        import os
        from tkinter import messagebox, filedialog
        import win32com.client as win32

        try:
            # Получение текущей директории и пути к Excel-файлу
            current_dir = os.getcwd()
            pstn_file_path = os.path.join(current_dir, 'PSTN.xlsx')

            if not os.path.exists(pstn_file_path):
                messagebox.showerror("Xəta", "PSTN.xlsx faylı tapılmadı")
                return



            # Запуск Excel
            excel = win32.Dispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False

            # Открываем рабочую книгу и нужный лист
            wb = excel.Workbooks.Open(pstn_file_path)
            sheet = wb.Sheets("Index (Output)")

            # Собираем данные: имя провайдера + критерии
            report_lines = []
            for row in range(2, 15):  # Строки 2–14
                pstn_name = sheet.Cells(row, 2).Value  # Столбец B
                if pstn_name:
                    criteria = []
                    for col in range(3, 24):  # Столбцы C до W (3–23)
                        cell_value = sheet.Cells(row, col).Value
                        if cell_value:
                            criteria.append(str(cell_value))
                    report_lines.append(f"{pstn_name}:\n  " + "\n  ".join(criteria))

            wb.Close(SaveChanges=False)

            if not report_lines:
                messagebox.showinfo("Məlumat", "PSTN faylında məlumat tapılmadı")
                return

            def get_pstn_date_from_excel():
                wb = load_workbook("PSTN.xlsx", data_only=True)  # Открываем файл с вычисленными значениями
                sheet = wb["Input"]  # Выбираем лист "Output"
                return sheet["T2"].value  # Возвращаем значение из ячейки E2

            # Создание PDF-файла (временный путь)
            from fpdf import FPDF

            # Создаём кастомный класс с автоматическим фоном на каждую страницу
            class PDFWithBackground(FPDF):
                def __init__(self, background_image, *args, **kwargs):
                    super().__init__(*args, **kwargs)
                    self.background_image = background_image
                def header(self):
                    # Устанавливаем фон на каждую страницу
                    self.image(self.background_image, x=0, y=0, w=210, h=297)

            # Создаём PDF с фоном
            background_image = os.path.join(current_dir, "Background_reports.jpg")
            if not os.path.exists(background_image):
                messagebox.showerror("Xəta", "Fayl Background_reports.jpg tapılmadı")
                return
            pdf = PDFWithBackground(background_image=background_image, format='A4', unit='mm')
            pdf.set_auto_page_break(auto=True, margin=15)

            # Добавляем шрифты
            pdf.add_font('DejaVu', '', 'DejaVuSans.ttf', uni=True)
            pdf.add_font('DejaVu', 'B', 'DejaVuSans-Bold.ttf', uni=True)
            pdf.add_page()

            # Добавляем название отчета
            pdf.set_font('DejaVu', 'B', 16)
            pdf.multi_cell(0, 10,
                           f'SABİT TELEFON OPERATORLARIN {get_pstn_date_from_excel()} RÜB VƏ İL ÜÇÜN TƏQDIM ETDIYI MƏLUMATLARLA BAĞLI XƏTALAR VƏ SƏBƏBLƏRI HAQQINDA HESABAT.',
                           0, 'C')

            # Отступ перед первым блоком
            y_position = pdf.get_y() + 10
            pdf.set_xy(10, y_position)

            # Перебираем блоки отчёта
            for block in report_lines:
                lines = block.strip().split("\n")
                if not lines:
                    continue

                provider_name = lines[0]
                criteria_lines = lines[1:]

                pdf.set_xy(10, y_position)
                pdf.set_font('DejaVu', 'B', 12)
                pdf.multi_cell(0, 10, txt=provider_name)
                y_position = pdf.get_y()

                pdf.set_font('DejaVu', '', 9)
                for idx, line in enumerate(criteria_lines, start=1):
                    formatted_line = f"{idx}. {line.strip()}"
                    pdf.multi_cell(0, 8, txt=formatted_line)
                    y_position = pdf.get_y()

                # Линия-разделитель
                pdf.set_line_width(0.5)
                pdf.line(10, y_position, 200, y_position)

                y_position += 5

                # Принудительная проверка на заполнение страницы
                if y_position > 270:
                    pdf.add_page()
                    y_position = 10




            # Диалог сохранения
            save_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                title="Yadda saxlayın",
                filetypes=[("PDF files", "*.pdf")],
                initialdir=os.getcwd(),
                initialfile="Report for PSTNs.pdf",
                confirmoverwrite=True
            )

            if save_path:
                pdf.output(save_path)
                messagebox.showinfo("Uğur", f"Hesabat yadda saxlanıldı")
            else:
                messagebox.showwarning("Xəbərdarlıq", "Hesabat yadda saxlanılmadı")

        except Exception as e:
            messagebox.showerror("Xəta", f"Xəta baş verdi: {str(e)}")

        finally:
            if 'excel' in locals():
                excel.Quit()

    def get_pstn_date_from_excel():
        wb = load_workbook("PSTN.xlsx", data_only=True)  # Открываем файл с вычисленными значениями
        sheet = wb["Input"]  # Выбираем лист "Output"
        return sheet["T2"].value  # Возвращаем значение из ячейки E2

    root = tk.Tk()
    root.title("PSTN-lərin yoxlanılması")
    root.iconbitmap("Icon.ico")
    root.geometry("400x250+400+170")
    root.resizable(False, False)

    # Загружаем изображение с помощью Pillow
    image = Image.open("Background.jpg")
    background_image = ImageTk.PhotoImage(image)

    # Canvas для фонового изображения
    canvas = tk.Canvas(root, width=450, height=250, highlightthickness=0)
    canvas.pack(fill="both", expand=True)

    # Устанавливаем фон
    canvas.create_image(0, 0, image=background_image, anchor="nw")

    # Добавляем заголовок (прозрачный фон)
    canvas.create_text(200, 50, text="PSTN-lərin yoxlanılması", font=("Helvetica", 22, "bold"), fill="black")

    # Список с позициями для кнопок: (x, y)
    button_positions = [(200, 120), (200, 190)]  # x, y для каждой кнопки

    # Размеры кнопок
    button_width = 25
    button_height = 2

    # Кнопка для генерации отчета
    ggenerate_button = tk.Button(root, text=f"Hesabat yaratmaq (dövr: {get_pstn_date_from_excel()})",
                                 font=("Arial", 14, "bold"),
                                 command=generate_report_pstn,
                                 bg="#40E0D0", fg="black", activebackground="#0000FF", activeforeground="white",
                                 relief="raised", width=button_width, height=button_height)
    canvas.create_window(button_positions[0][0], button_positions[0][1], window=ggenerate_button)

    # Кнопка для поиска провайдера
    search_button = tk.Button(root, text=f"PSTN üzrə axtarış (dövr: {get_pstn_date_from_excel()})",
                              font=("Arial", 14, "bold"),
                              command=search_pstn,
                              bg="#7FFFD4", fg="black", activebackground="#0000FF", activeforeground="white",
                              relief="raised", width=button_width, height=button_height)
    canvas.create_window(button_positions[1][0], button_positions[1][1], window=search_button)

    root.mainloop()
    dcstool_main()




def dcstool_main():
    # Код для основного окна (DCS Tool (PSTN and ISP))
    from tkinter import messagebox
    import tkinter as tk
    from PIL import Image, ImageTk

    # Функции кнопок
    def on_button0_click():
        get_quarter_input(update_database_isp_pstn, root)


    def on_button1_click():
        root.destroy()  # Закрыть текущее окно
        dcs_tool_isp()  # Вызвать функцию dcs_tool_isp

    def on_button2_click():
        root.destroy()  # Закрыть текущее окно
        dcs_tool_pstn()  # Вызвать функцию dcs_tool_pstn

    def on_button3_click():
        messagebox.showinfo("Mobillər", "Tezliklə...")



    # Создаем главное окно
    import os
    import sys

    # Список конкретных файлов Excel с пробелами в названиях
    required_files = ["ISP.xlsx", "PSTN.xlsx", "Icon.ico", "Background.jpg"]

    # Получаем текущую рабочую директорию (папка с программой)
    current_directory = os.getcwd()

    # Проверка наличия файлов
    for file_name in required_files:
        file_path = os.path.join(current_directory, file_name)  # Путь к файлу
        if not os.path.isfile(file_path):  # Проверка наличия файла
            # Создаём окно для отображения ошибки
            root = tk.Tk()
            root.withdraw()  # Скрываем основное окно
            messagebox.showerror("Xəta", "Tələb olunan fayllar proqram qovluğunda yoxdur!")
            sys.exit(1)  # Завершаем программу с кодом ошибки

    root = tk.Tk()
    root.title("DCS")
    root.geometry("450x450+400+170")
    root.iconbitmap("Icon.ico")
    root.resizable(False, False)  # Окно нельзя изменять по размеру


    # Загружаем изображение с помощью Pillow
    image = Image.open("Background.jpg")
    background_image = ImageTk.PhotoImage(image)

    # Canvas для фонового изображения
    canvas = tk.Canvas(root, width=450, height=450, highlightthickness=0)
    canvas.pack(fill="both", expand=True)

    # Устанавливаем фон
    canvas.create_image(0, 0, image=background_image, anchor="nw")

    # Добавляем заголовок (прозрачный фон)
    canvas.create_text(225, 50, text="Təsərrüfat DCS yoxlanması", font=("Helvetica", 22, "bold"), fill="black")

    # Координаты кнопок (по центру экрана)
    button_x = 225  # Все кнопки выровнены по центру
    button_y_positions = [120, 180, 240, 300]  # Отступы между кнопками

    # Размеры кнопок
    button_width = 20  # Ширина кнопки для остальных
    button_height = 2  # Высота кнопки для остальных

    # Размеры для database_button
    database_button_width = 25  # Ширина кнопки для database_button
    database_button_height = 1  # Высота кнопки для database_button

    # Добавим расстояние между кнопками
    button_spacing = 10  # Расстояние между кнопками

    # Создаем кнопки с одинаковыми размерами и пробелами между ними
    database_button = tk.Button(root, text="Məlumat Bazasının Yenilənməsi", font=("Arial", 14, "bold"),
                                command=on_button0_click,
                                bg="#008080", fg="white", activebackground="#32CD32", activeforeground="white",
                                relief="raised", width=database_button_width, height=database_button_height)
    canvas.create_window(button_x, button_y_positions[0], window=database_button)

    isp_button = tk.Button(root, text="ISP-lərin yoxlanılması",
                           font=("Arial", 14, "bold"),
                           command=on_button1_click,
                           bg="#40E0D0", fg="black", activebackground="#0000FF", activeforeground="white",
                           relief="raised", width=button_width, height=button_height)
    canvas.create_window(button_x, button_y_positions[1] + button_spacing, window=isp_button)

    pstn_button = tk.Button(root, text="PSTN-lərin yoxlanılması",
                            font=("Arial", 14, "bold"),
                            command=on_button2_click,
                            bg="#7FFFD4", fg="black", activebackground="#0000FF", activeforeground="white",
                            relief="raised", width=button_width, height=button_height)
    canvas.create_window(button_x, button_y_positions[2] + 2 * button_spacing, window=pstn_button)

    mobiles_button = tk.Button(root, text="Mobillərin yoxlanılması", font=("Arial", 14, "bold"),
                               command=on_button3_click,
                               bg="#87CEEB", fg="black", activebackground="#0000FF", activeforeground="white",
                               relief="raised", width=button_width, height=button_height)
    canvas.create_window(button_x, button_y_positions[3] + 3 * button_spacing, window=mobiles_button)

    # Запускаем главный цикл
    root.mainloop()

dcstool_main()

